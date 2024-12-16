from bisect import bisect_left, bisect_right
from openpyxl import load_workbook
from datetime import datetime


def load_excel(file_path):
    """
    if the excel file isnt sorted by dates, we won't be able to binary search it
    """
    wb = load_workbook(file_path)  # load workbook
    ws = wb.active  # get the active worksheet

    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    date_col_idx = headers.index("Date")  # find the index of the "Date" column
    category_col_idx = headers.index("Category")

    data = []

    # parse rows and ensure dates are in datetime format
    for row in rows[1:]:
        try:
            # convert date -> datetime
            date = row[date_col_idx]
            if isinstance(date, str):
                date = datetime.strptime(date, '%m/%d/%Y')
            elif isinstance(date, datetime):
                pass
            else:
                date = None  # handle cases where the date is None or invalid

            # only append rows with valid dates
            if date is not None:
                data.append(row)
        except Exception as e:
            print(f"Skipping row due to error: {e}")


    data.sort(key=lambda x: x[date_col_idx] if x[date_col_idx] else datetime.min)

    return wb, headers, data, ws  # Return wb (workbook) so it can be saved later


def find_date_range_indices(data, start_date, end_date, date_col_idx):
    """
    binary search
    """
    start_index = bisect_left([row[date_col_idx] for row in data], start_date)
    end_index = bisect_right([row[date_col_idx] for row in data], end_date) - 1
    return start_index, end_index


def filter_by_category(data, start_date, end_date, category, headers):
    """
    filter the rows using user input
    """
    date_col_idx = headers.index("Date")
    category_col_idx = headers.index("Category")

    start_index, end_index = find_date_range_indices(data, start_date, end_date, date_col_idx)


    filtered_data = [row for row in data[start_index:end_index + 1] if row[category_col_idx] == category]

    return filtered_data


def write_to_excel(file_path, filtered_rows, start_col, start_row, headers, wb, ws):
    """
    write the extracted data onto the excel sheet itself
    """
    relevant_columns = ['Date', 'Tramsaction', 'Notes', 'Category', 'Check #', 'Amount', 'Deposit']

    # determine the starting column index (e.g., "J" -> 10)
    start_col_index = ord(start_col.upper()) - ord('A') + 1

    # use relevant columns as to avoid the script pulling from previous iterations of the same script
    relevant_indices = [headers.index(col) for col in relevant_columns]
    filtered_rows = [
        tuple(row[idx] for idx in relevant_indices) for row in filtered_rows
    ]

    # debugging print
    print("Filtered Rows to be Written:")
    for row in filtered_rows:
        print(row)

    # calculate the area of the area to be cleared
    num_rows_to_clear = len(filtered_rows) + 1
    num_cols_to_clear = len(relevant_columns)

    # clear only the necessary area (below the current data to avoid overwriting)
    for row in range(start_row, start_row + num_rows_to_clear):
        for col in range(start_col_index, start_col_index + num_cols_to_clear):
            ws.cell(row=row, column=col).value = None

    # write the headers in the starting column and row
    for col_idx, col_name in enumerate(relevant_columns, start=start_col_index):
        ws.cell(row=start_row, column=col_idx).value = col_name

    # write data rows under the headers
    for row_idx, row in enumerate(filtered_rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col_index):
            # Format date without time if value is a date
            if isinstance(value, datetime):
                value = value.strftime('%m/%d/%Y')  # Format as MM/DD/YYYY
            ws.cell(row=row_idx, column=col_idx).value = value

    wb.save(file_path)
    print(f"Filtered rows written to {file_path} starting from column {start_col}{start_row}")


def main():
    # prompt the user for file path, date range, and category
    file_path = input("Enter the path to your Excel file: ").strip()
    start_date = datetime.strptime(input("Enter the start date (MM/DD/YYYY): ").strip(), "%m/%d/%Y")
    end_date = datetime.strptime(input("Enter the end date (MM/DD/YYYY): ").strip(), "%m/%d/%Y")
    category = input("Enter the category to filter by: ").strip()

    # ask the user where to write the output in the excel sheet
    start_col = input("Enter the starting column for the output (e.g., 'J'): ").strip()
    start_row = int(input("Enter the starting row for the output (e.g., '1'): ").strip())


    try:
        wb, headers, data, ws = load_excel(file_path)  # Get the workbook (wb)
        filtered_rows = filter_by_category(data, start_date, end_date, category, headers)

        if not filtered_rows:
            print(f"No rows found for category '{category}' between {start_date} and {end_date}.")
        else:
            print(f"Filtered rows for category '{category}' between {start_date} and {end_date}:")
            for row in filtered_rows:
                print(row)

            # write
            write_to_excel(file_path, filtered_rows, start_col, start_row, headers, wb, ws)

    except Exception as e:
        print(f"An error occurred: {e}")

    input("Press Enter to exit...")


if __name__ == "__main__":
    main()
